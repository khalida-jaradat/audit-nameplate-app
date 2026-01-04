import os
import re
import io
import json
import base64
from datetime import datetime

import streamlit as st
import pandas as pd
from PIL import Image

import requests

# Optional OCR deps (won't crash if missing)
try:
    import cv2
    import numpy as np
    import pytesseract
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# =========================================================
# 0) CONFIG
# =========================================================

AUDIT_TEMPLATES = {
    "Hotels": [
        "Lobby",
        "Rooms",
        "Kitchen",
        "Pool",
        "Boiler system",
        "Electrical system",
        "Generator system",
        "Laundry",
    ],
    "Factories": [
        "Production Lines",
        "Workshop / Maintenance",
        "Warehouse",
        "Compressor system",
        "Boiler system",
        "Electrical system",
        "Generator Room",
        "Office Area",
        "Outdoor / Yard",
        "Chiller Plant",
        "Chilled Water Network",
        "Compressed Air Network",
        "Hot Water Network",
        "Steam Network",
        "Steam Generator",
    ],
    "Schools": [
        "Classrooms",
        "Computer Labs",
        "Science Labs",
        "Administration",
        "Library",
        "Cafeteria",
        "Sports Hall / Outdoor Area",
        "Electrical system",
        "Generator system",
    ],
    "Hospitals": [
        "ER",
        "ICU",
        "Wards",
        "Operating Theatres",
        "Labs",
        "Imaging Area",
        "Pharmacy",
        "Kitchen",
        "Laundry",
        "Boiler system",
        "Electrical system",
        "Generator system",
    ],
}

OUTPUT_DIR = "outputs"
CSV_PATH = os.path.join(OUTPUT_DIR, "audit_nameplate_records.csv")
THUMBS_DIR = os.path.join(OUTPUT_DIR, "_thumbs")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(THUMBS_DIR, exist_ok=True)

DEFAULT_TESS_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"  # for local Windows only


# =========================================================
# 1) HELPERS
# =========================================================

def safe_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^A-Za-z0-9_\- ]+", "", s).strip()
    return s.replace(" ", "_") if s else "AUDIT"


def append_record(row: dict):
    if os.path.exists(CSV_PATH):
        df = pd.read_csv(CSV_PATH)
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    else:
        df = pd.DataFrame([row])
    df.to_csv(CSV_PATH, index=False)


def create_thumbnail(img_path: str) -> str | None:
    try:
        if not img_path or not os.path.exists(img_path):
            return None
        base = os.path.basename(img_path)
        name, _ = os.path.splitext(base)
        thumb_path = os.path.join(THUMBS_DIR, f"{name}_thumb.png")

        pil = Image.open(img_path).convert("RGB")
        pil.thumbnail((260, 160))
        pil.save(thumb_path, format="PNG")
        return thumb_path
    except Exception:
        return None


# =========================================================
# 2) CLASSIC OCR (optional)
# =========================================================

def ensure_tesseract():
    if not OCR_AVAILABLE:
        return
    env_cmd = os.environ.get("TESSERACT_CMD", "").strip()
    if env_cmd and os.path.exists(env_cmd):
        pytesseract.pytesseract.tesseract_cmd = env_cmd
        return
    if os.name == "nt" and os.path.exists(DEFAULT_TESS_CMD):
        pytesseract.pytesseract.tesseract_cmd = DEFAULT_TESS_CMD


def preprocess_for_ocr(pil_img: Image.Image):
    img = np.array(pil_img.convert("RGB"))
    img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    img = cv2.resize(img, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)

    th = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31, 10
    )
    return th


def classic_ocr_text(pil_img: Image.Image) -> str:
    if not OCR_AVAILABLE:
        return ""
    try:
        ensure_tesseract()
        proc = preprocess_for_ocr(pil_img)
        text = pytesseract.image_to_string(proc, config="--oem 3 --psm 6", lang="eng")
        return text or ""
    except Exception:
        return ""


# =========================================================
# 3) GEMINI VISION
# =========================================================

def get_gemini_settings():
    api_key = None
    model = "gemini-2.5-flash"
    try:
        api_key = st.secrets.get("GEMINI_API_KEY", None)
        model = st.secrets.get("GEMINI_MODEL", model)
    except Exception:
        pass

    if not api_key:
        api_key = os.environ.get("GEMINI_API_KEY")

    model = (model or "gemini-2.5-flash").strip()
    model = model.replace("models/", "").strip()

    return api_key, model


def extract_first_json(text: str) -> str:
    if not text:
        return ""
    t = text.strip()

    # remove common fences
    t = t.replace("```json", "").replace("```", "").strip()

    # try direct
    try:
        json.loads(t)
        return t
    except Exception:
        pass

    # find a JSON object inside
    m = re.search(r"\{.*\}", t, flags=re.S)
    if m:
        return m.group(0).strip()

    return ""


def ai_extract_fields(pil_img: Image.Image) -> tuple[str, dict]:
    """
    Gemini Vision عبر REST (بدون مكتبات إضافية).
    يرجّع:
      - raw_text: النص الخام (JSON أو رسالة خطأ) للـ debug
      - fields: dict جاهز للعرض (Model/Serial/Voltage_V/Current_A/Power_kW/Frequency_Hz)
    """
    empty = {
        "Model": "",
        "Serial": "",
        "Voltage_V": "",
        "Current_A": "",
        "Power_kW": "",
        "Frequency_Hz": "",
    }

    # 1) اقرأ المفتاح والموديل من secrets أو env
    api_key = ""
    model = "gemini-2.5-flash"
    try:
        api_key = (st.secrets.get("GEMINI_API_KEY", "") or "").strip()
        model = (st.secrets.get("GEMINI_MODEL", model) or model).strip()
    except Exception:
        api_key = (os.environ.get("GEMINI_API_KEY", "") or "").strip()
        model = (os.environ.get("GEMINI_MODEL", model) or model).strip()

    if not api_key:
        return "Gemini API key missing: set GEMINI_API_KEY in Streamlit Secrets.", empty

    # 2) حضّر الصورة (JPEG) + base64
    try:
        img = pil_img.convert("RGB")
        # تصغير بسيط عشان القراءة تكون أوضح وأسرع (اختياري)
        max_w = 1400
        if img.width > max_w:
            ratio = max_w / float(img.width)
            img = img.resize((max_w, int(img.height * ratio)))

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    except Exception as e:
        return f"Image encode error: {e}", empty

    # 3) Prompt (شدّدنا على: JSON فقط + بدون أسطر داخل القيم)
    prompt = (
        "You are an expert electrical auditor. Extract nameplate fields from the image.\n"
        "Return ONLY valid JSON (no markdown, no code fences).\n"
        "Rules:\n"
        "- All string values must be SINGLE-LINE (no line breaks).\n"
        "- If missing, use null.\n"
        "- voltage_v: number (if range like 220-240, return 230).\n"
        "- power_kw: number in kW (if label shows W, convert to kW).\n\n"
        "Return exactly this schema:\n"
        "{\n"
        '  "model": string|null,\n'
        '  "serial": string|null,\n'
        '  "voltage_v": number|null,\n'
        '  "current_a": number|null,\n'
        '  "power_kw": number|null,\n'
        '  "frequency_hz": number|null\n'
        "}\n"
    )

    # 4) Gemini REST call مع JSON mode
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": "image/jpeg", "data": b64}},
                ],
            }
        ],
        "generationConfig": {
            "responseMimeType": "application/json",
            "temperature": 0.0,
            "maxOutputTokens": 220,
        },
    }

    try:
        r = requests.post(url, json=payload, timeout=60)
        if not r.ok:
            return f"Gemini HTTP {r.status_code}: {r.text[:1200]}", empty
        resp = r.json()
    except Exception as e:
        return f"Gemini request error: {e}", empty

    # 5) اقرأ النص المرجّع
    raw_text = ""
    try:
        raw_text = resp["candidates"][0]["content"]["parts"][0]["text"]
    except Exception:
        raw_text = json.dumps(resp, ensure_ascii=False)[:2000]

    # 6) Parser أقوى: حمّلي JSON مباشرة، وإذا فشل قصّي { ... }
    def _strip_fences(s: str) -> str:
        s = (s or "").strip()
        s = re.sub(r"^```(?:json)?\s*", "", s, flags=re.IGNORECASE)
        s = re.sub(r"\s*```$", "", s)
        return s.strip()

    cleaned = _strip_fences(raw_text)

    obj = None
    try:
        obj = json.loads(cleaned)
    except Exception:
        m = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if m:
            try:
                obj = json.loads(m.group(0))
            except Exception:
                obj = None

    if not isinstance(obj, dict):
        return f"Gemini returned no JSON. Raw: {raw_text[:1200]}", empty

    # 7) Normalize + safeguards
    def to_num(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        # التقط أول رقم (يدعم 220-240)
        rng = re.findall(r"[-+]?\d+(\.\d+)?", s)
        if not rng:
            return None
        # إذا كان Range (220-240) خذ الوسط
        nums = [float(re.findall(r"[-+]?\d+(\.\d+)?", t)[0]) if False else float(re.search(r"[-+]?\d+(\.\d+)?", t).group(0)) for t in re.findall(r"[-+]?\d+(\.\d+)?", s)]
        # ^ السطر أعلاه حساس؛ الأسهل:
        nums2 = [float(m.group(0)) for m in re.finditer(r"[-+]?\d+(\.\d+)?", s)]
        if len(nums2) >= 2 and "-" in s:
            return (nums2[0] + nums2[1]) / 2.0
        return nums2[0] if nums2 else None

    model_val = obj.get("model")
    serial_val = obj.get("serial")

    v = to_num(obj.get("voltage_v"))
    a = to_num(obj.get("current_a"))
    p = to_num(obj.get("power_kw"))
    hz = to_num(obj.get("frequency_hz"))

    # لو power طلع كبير جدًا بالغلط (W)، نحاول نحوله لـ kW
    if p is not None and p > 50:  # غالبًا W
        p = p / 1000.0

    fields = {
        "Model": (str(model_val).replace("\n", " ").strip() if model_val else ""),
        "Serial": (str(serial_val).replace("\n", " ").strip() if serial_val else ""),
        "Voltage_V": ("" if v is None else str(round(v, 3)).rstrip("0").rstrip(".")),
        "Current_A": ("" if a is None else str(round(a, 3)).rstrip("0").rstrip(".")),
        "Power_kW": ("" if p is None else str(round(p, 3)).rstrip("0").rstrip(".")),
        "Frequency_Hz": ("" if hz is None else str(round(hz, 3)).rstrip("0").rstrip(".")),
    }

    return json.dumps(obj, ensure_ascii=False), fields


def analyze_nameplate(pil_img: Image.Image) -> tuple[str, dict]:
    classic = classic_ocr_text(pil_img)  # optional
    ai_json, ai_fields = ai_extract_fields(pil_img)

    raw_combined = ""
    if classic:
        raw_combined += classic + "\n\n---\n\n"
    if ai_json:
        raw_combined += ai_json

    return raw_combined, ai_fields


# =========================================================
# 4) EXCEL EXPORT (with thumbnails)
# =========================================================

def build_excel_report(df: pd.DataFrame, out_path: str):
    cols = [
        "Record ID",
        "Place Name",
        "PlaceIndex",
        "Model",
        "Serial",
        "Voltage (V)",
        "Current (A)",
        "Power (kW)",
        "Frequency (Hz)",
        "Capture DateTime",
        "Notes",
        "Nameplate Image",
    ]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center", wrap_text=True)

    def set_col_widths(ws):
        widths = {
            1: 10,
            2: 24,
            3: 10,
            4: 20,
            5: 22,
            6: 12,
            7: 12,
            8: 12,
            9: 14,
            10: 22,
            11: 22,
            12: 30,
        }
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def add_headers(ws):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    wb = Workbook()
    wb.remove(wb.active)

    if df.empty:
        wb.save(out_path)
        return

    # group sheets by Place (base)
    places = sorted(df["Place"].dropna().unique().tolist())
    rec_counter = 1

    for place in places:
        ws = wb.create_sheet(title=str(place)[:31])
        add_headers(ws)
        set_col_widths(ws)
        ws.freeze_panes = "A2"

        df_place = df[df["Place"] == place].copy()

        for _, row in df_place.iterrows():
            record_id = f"{rec_counter:03d}"
            rec_counter += 1

            place_index = row.get("PlaceIndex", "")
            area_name = str(row.get("AreaName", "") or "").strip()
            place_label = place if not area_name else f"{place} - {area_name}"

            model = row.get("Model", "")
            serial = row.get("Serial", "")
            voltage = row.get("Voltage_V", "")
            current = row.get("Current_A", "")
            power = row.get("Power_kW", "")
            freq = row.get("Frequency_Hz", "")
            ts = row.get("Timestamp", "")
            notes = row.get("Notes", "")
            img_path = str(row.get("ImagePath", "")).strip()

            ws.append(
                [
                    record_id,
                    place_label,
                    place_index,
                    model,
                    serial,
                    voltage,
                    current,
                    power,
                    freq,
                    ts,
                    notes,
                    "",
                ]
            )

            r = ws.max_row
            ws.row_dimensions[r].height = 120

            img_path = os.path.normpath(img_path)
            thumb_path = create_thumbnail(img_path)
            if thumb_path and os.path.exists(thumb_path):
                try:
                    xl_img = XLImage(thumb_path)
                    anchor_cell = f"{get_column_letter(len(cols))}{r}"
                    xl_img.anchor = anchor_cell
                    ws.add_image(xl_img)
                except Exception:
                    pass

            for c in range(1, len(cols)):
                ws.cell(row=r, column=c).alignment = center

    wb.save(out_path)


# =========================================================
# 5) STREAMLIT UI
# =========================================================

st.set_page_config(page_title="Audit Nameplate (Gemini)", layout="wide")
st.title("Audit Nameplate OCR / Vision (Gemini)")

api_key, model_name = get_gemini_settings()

with st.sidebar:
    st.subheader("System Status")
    st.write(f"Model: `google/{model_name}`")
    if api_key:
        st.success("AI Vision: ENABLED ✅")
        st.write(f"Key length: {len(api_key)}")
    else:
        st.error("AI Vision: DISABLED ❌ (Missing GEMINI_API_KEY)")

    if not OCR_AVAILABLE:
        st.info("Classic OCR: disabled (missing cv2/pytesseract). OK for phone usage.")
    else:
        st.write("Classic OCR: optional ✅")


st.write(
    "Workflow: select audit type → facility → area → capture/upload nameplate → AI extracts fields → save → export Excel."
)

col1, col2 = st.columns(2)
with col1:
    audit_type = st.selectbox("Audit Type", list(AUDIT_TEMPLATES.keys()))
with col2:
    facility_name = st.text_input("Facility Name", placeholder="e.g., Demo Factory")

default_places = AUDIT_TEMPLATES.get(audit_type, [])
default_places = [p.strip() for p in default_places if p.strip()]

st.divider()
st.subheader("Audit Components (editable list)")
places_text = st.text_area(
    "One place per line",
    value="\n".join(default_places),
    height=180,
)
places = [p.strip() for p in places_text.splitlines() if p.strip()]

if not places:
    st.warning("Please add at least one place (one per line).")

st.divider()
st.subheader("Choose a place to capture nameplates")

place = st.selectbox("Place", places) if places else None
count = st.number_input(
    "How many instances of this place?",
    min_value=1,
    max_value=20,
    value=1,
)

# optional custom area name (stored next to place in Excel)
area_name = st.text_input(
    "Custom area name (optional) – e.g., Main Kitchen, West Lobby",
    value="",
)

if facility_name and place:
    st.info(f"You will capture nameplates for: **{facility_name} → {place} (1..{int(count)})**")

st.divider()
st.subheader("Capture nameplates")

def open_image_any(file_obj) -> Image.Image | None:
    """
    Supports normal images. HEIC/HEIF needs extra libs; if not available,
    user should upload JPG/PNG.
    """
    try:
        return Image.open(file_obj)
    except Exception:
        # Try HEIC via pillow-heif if installed
        try:
            import pillow_heif  # type: ignore
            pillow_heif.register_heif_opener()
            return Image.open(file_obj)
        except Exception:
            return None

if facility_name and place:
    for i in range(1, int(count) + 1):
        with st.expander(f"{place} #{i}", expanded=(i == 1)):
            camera_photo = st.camera_input(
                f"Take photo for {place} #{i}",
                key=f"cam_{audit_type}_{place}_{i}",
            )
            uploaded = st.file_uploader(
                f"Or upload existing image for {place} #{i}",
                type=["png", "jpg", "jpeg", "webp", "heic", "heif"],
                key=f"upl_{audit_type}_{place}_{i}",
            )

            image_file = camera_photo or uploaded

            if image_file:
                pil_img = open_image_any(image_file)
                if pil_img is None:
                    st.error("Can't open this image format. Please upload JPG/PNG (or enable HEIC support).")
                    st.stop()

                st.image(pil_img, caption="Uploaded image", use_container_width=True)

                with st.spinner("Analyzing image (Gemini Vision)..."):
                    raw, fields = analyze_nameplate(pil_img)

                st.markdown("### Extracted fields (edit if needed)")
                c1, c2, c3, c4, c5, c6 = st.columns(6)
                with c1:
                    model_val = st.text_input("Model", value=fields.get("Model") or "", key=f"model_{place}_{i}")
                with c2:
                    serial_val = st.text_input("Serial", value=fields.get("Serial") or "", key=f"serial_{place}_{i}")
                with c3:
                    voltage_val = st.text_input("Voltage (V)", value=fields.get("Voltage_V") or "", key=f"volt_{place}_{i}")
                with c4:
                    current_val = st.text_input("Current (A)", value=fields.get("Current_A") or "", key=f"curr_{place}_{i}")
                with c5:
                    power_val = st.text_input("Power (kW)", value=fields.get("Power_kW") or "", key=f"pwr_{place}_{i}")
                with c6:
                    freq_val = st.text_input("Frequency (Hz)", value=fields.get("Frequency_Hz") or "", key=f"hz_{place}_{i}")

                notes_val = st.text_input("Notes (optional)", value="", key=f"notes_{place}_{i}")

                with st.expander("Raw OCR / AI output (debug)"):
                    st.code(raw[:4000] if raw else "")

                if st.button(f"Save record for {place} #{i}", key=f"save_{place}_{i}"):
                    safe_fac = safe_name(facility_name)
                    safe_place = safe_name(place)

                    dir_path = os.path.join(OUTPUT_DIR, safe_fac, f"{safe_place}_{i}")
                    os.makedirs(dir_path, exist_ok=True)

                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    img_path = os.path.join(dir_path, f"nameplate_{ts}.png")
                    pil_img.convert("RGB").save(img_path)
                    img_path = os.path.abspath(img_path)

                    row = {
                        "Timestamp": datetime.now().isoformat(timespec="seconds"),
                        "AuditType": audit_type,
                        "Facility": facility_name,
                        "Place": place,                 # base place (sheet grouping)
                        "AreaName": area_name.strip(),  # optional custom name
                        "PlaceIndex": i,
                        "Model": model_val,
                        "Serial": serial_val,
                        "Voltage_V": voltage_val,
                        "Current_A": current_val,
                        "Power_kW": power_val,
                        "Frequency_Hz": freq_val,
                        "Notes": notes_val,
                        "ImagePath": img_path,
                        "RawOCR": raw,
                        "AIModel": model_name,
                    }

                    append_record(row)

                    json_path = os.path.join(dir_path, f"record_{ts}.json")
                    with open(json_path, "w", encoding="utf-8") as f:
                        json.dump(row, f, ensure_ascii=False, indent=2)

                    st.success("Saved! Record added to CSV with local image.")


# =========================================================
# 6) CURRENT CSV PREVIEW (filtered by facility)
# =========================================================

st.divider()
st.subheader("Current CSV preview")

if os.path.exists(CSV_PATH):
    df_all = pd.read_csv(CSV_PATH)

    if facility_name:
        df_show = df_all[df_all["Facility"] == facility_name].copy()
    else:
        df_show = df_all.copy()

    st.dataframe(df_show.tail(50), use_container_width=True)
    st.caption(f"Saved at: {CSV_PATH}")
else:
    st.write("No records yet.")


# =========================================================
# 7) EXPORT EXCEL
# =========================================================

st.divider()
st.subheader("Export Excel (embedded images by Place)")

if os.path.exists(CSV_PATH):
    df_all = pd.read_csv(CSV_PATH)

    if facility_name:
        df_use = df_all[df_all["Facility"] == facility_name].copy()
    else:
        df_use = df_all.copy()

    if not df_use.empty:
        safe_fac = safe_name(facility_name or "AUDIT")
        xlsx_path = os.path.join(OUTPUT_DIR, f"AUDIT_{safe_fac}_{datetime.now().strftime('%Y%m%d')}.xlsx")

        if st.button("Generate Excel with embedded images"):
            for col in ["Notes", "RawOCR", "Current_A", "AreaName"]:
                if col not in df_use.columns:
                    df_use[col] = ""

            build_excel_report(df_use, xlsx_path)
            st.success("Excel generated successfully!")

            with open(xlsx_path, "rb") as f:
                st.download_button(
                    "Download Excel Report",
                    data=f,
                    file_name=os.path.basename(xlsx_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    else:
        st.info("No records for this facility yet. Save at least one nameplate record.")
else:
    st.info("No CSV records yet. Save at least one nameplate record first.")

